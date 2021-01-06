# -*- coding: utf-8 -*-
"""
Created on Mon Sep  9 17:05:43 2019

@author: Reuben
"""

import json
import importlib
from . import utils

def is_importable(module_name):
    spec = importlib.util.find_spec(module_name)
    return spec is not None

IS_YAML = is_importable('ruamel.yaml')
if IS_YAML:
    import ruamel.yaml as yml

IS_XLSX = is_importable('openpyxl')
if IS_XLSX:
    from openpyxl import load_workbook
  


class Manager():
    default_handler = 'yaml'
    
    def __init__(self):
        handlers = {}
        if IS_YAML:
            handlers['yaml'] = Yaml()
        if IS_XLSX:
            handlers['xlsx'] = Xlsx()
        self.handlers = handlers
        self.specified = None
        
    def add_handler(self, key, handler):
        self.handlers[key] = handler
        
    def specify(self, key):
        self.specified = key
        
    def load(self, source, handler=None, **kwargs):
        if handler is not None:
            return self.handlers[handler].load(source, **kwargs)
        if self.specified is not None:
            return self.handlers[self.specified].load(source, **kwargs)
        for k, handler in self.handlers.items():
            if handler.suitable(source, **kwargs):
                return handler.load(source, **kwargs)
        f = source + '.yaml'
        return self.handlers[self.default_handler].load(f, **kwargs)

    def save(self, dct, target, handler=None, **kwargs):
        if handler is not None:
            return self.handlers[handler].save(target, **kwargs)
        if self.specified is not None:
            return self.handlers[self.specified].save(dct, target, **kwargs)
        for k, handler in self.handlers.items():
            if handler.suitable(target, **kwargs):
                return handler.save(dct, target, **kwargs)
        f = target + '.yaml'
        return self.handlers[self.default_handler].save(dct, f, **kwargs)
        

class Handler():
    def __init__(self):
        if not is_importable(self._import_name):
            raise ImportError("Package '" + self._import_name 
                              + "' is required for class " + str(type(self)))

    def suitable(self, fname, **kwargs):
        passes = [fname.lower().endswith('.' + e.lower()) 
                  for e in self._valid_suffixes]
        return any(passes)
    

class Yaml(Handler):
    _import_name = 'ruamel.yaml'    
    _valid_suffixes = ['yml', 'yaml']
    
    def load(self, fname, **kwargs):
        with open(fname, mode='r') as f:
            dct = yml.safe_load(f)
        return dct
    
    def save(self, dct, fname, **kwargs):
        with open(fname, mode='w') as f:
            y = yml.YAML()
            y.dump(dict(dct), f)
            

class Xlsx(Handler):
    _import_name = 'openpyxl'    
    _valid_suffixes = ['xlsx']
    
    
    def _unpack(self, obj):
        lst = []
        if isinstance(obj, (list, tuple)):
            for subobj in obj:
                lst.extend(self._unpack(subobj))
        else:
            lst.append(obj)
        return lst
    
    def _get_cells(self, wb, range_name):
        named_range = wb.defined_names[range_name]
        cells = []
        for title, coord in named_range.destinations:
            ws = wb[title]
            tups = ws[coord]
            lst = self._unpack(tups)
            cells.extend(lst)
        return cells
    
    def _check_rows(self, wb, key_cells, value_cells):
        if len(key_cells) != len(value_cells):
            raise KeyError("The number of value cells does not match the "
                           + "number of key cells.")
        
        for key_cell, value_cell in zip(key_cells, value_cells):
            if key_cell.row != value_cell.row:
                raise ValueError("The value cells and keys do not align. " +
                    "The rows for keys must match the rows for values.")
            if key_cell.parent != value_cell.parent:
                raise ValueError("The value cells and keys do not align. " +
                "The rows for keys and values must be on the same sheet.")
    
    def _read_vals(self, wb, key_cells, value_cells):
        dct = {}
        for key_cell, value_cell in zip(key_cells, value_cells):
            value = value_cell.value
            if value is None:
                continue
            if isinstance(value, (float, int)):
                processed_value = value
            if isinstance(value, str):
                if value[0] in ['"', "'", '[', '{']:
                    processed_value = json.loads(value.replace('\'', '"'))
                else:
                    processed_value = value
            dct[key_cell.value] = processed_value
        return dct
    
    def _write_vals(self, dct, wb, key_cells, value_cells):
        for key_cell, value_cell in zip(key_cells, value_cells):
            key = key_cell.value
            if not key in dct:
                continue
            value = utils.denumpify(dct[key])
            if value is None:
                continue
            if isinstance(value, (str, float, int)):
                write_value = value
            elif isinstance(value, (list, dict)):
                write_value = json.dumps(value)
            value_cell.value = write_value
    
    def _get_cell_pairs(self, wb, keys, values):
        key_cells = self._get_cells(wb, keys)
        value_cells = self._get_cells(wb, values)
        self._check_rows(wb, key_cells, value_cells)
        return key_cells, value_cells
    
    def load(self, fname, keys='dotkeys', values='values', **kwargs):
        wb = load_workbook(filename=fname)
        key_cells, value_cells = self._get_cell_pairs(wb, keys, values)
        dct = self._read_vals(wb, key_cells, value_cells)
        return dct
    
    def save(self, dct, fname, loadname, keys='dotkeys', values='values',
             **kwargs):
        wb = load_workbook(filename=loadname)
        key_cells, value_cells = self._get_cell_pairs(wb, keys, values)
        self._write_vals(dct, wb, key_cells, value_cells)
        wb.save(fname)

            
manager = Manager()
load = manager.load
save = manager.save
